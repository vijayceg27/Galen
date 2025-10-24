"""
Survey Data Crosstab Generator

This script reads survey data from an Excel workbook and generates crosstabs
with percentages and sample sizes for each question, segmented by key variables.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import shutil
import os
from datetime import datetime
warnings.filterwarnings('ignore')


class CrosstabGenerator:
    """Generate crosstabs from survey data with segmentation."""
    
    def __init__(self, file_path, output_path=None):
        """
        Initialize the crosstab generator.
        
        Args:
            file_path: Path to the source Excel workbook
            output_path: Path for the output file (optional, will be auto-generated if not provided)
        """
        self.source_file_path = file_path
        self.output_file_path = output_path
        self.file_path = None  # Will be set after copying
        self.data_df = None
        self.layout_df = None
        self.segment_questions = ['SAMPLE_TYPE', 'S2', 'SETTING']
        
    def load_data(self):
        """Load Data and Layout tabs from the Excel file."""
        print("Loading data from Excel file...")
        
        # Load from SOURCE file (before copying) to avoid filter issues
        source_file = self.source_file_path
        
        # Try loading with pandas first
        try:
            self.data_df = pd.read_excel(source_file, sheet_name='Data', engine='openpyxl')
            self.layout_df = pd.read_excel(source_file, sheet_name='Layout', engine='openpyxl')
        except (ValueError, Exception) as e:
            if "wildcard" in str(e) or "filter" in str(e).lower():
                print("  Note: Excel file has filters/formatting issues, using alternative loading method...")
                # Use xlrd or manual loading
                try:
                    # Try with xlrd engine (older Excel format)
                    self.data_df = pd.read_excel(source_file, sheet_name='Data', engine='xlrd')
                    self.layout_df = pd.read_excel(source_file, sheet_name='Layout', engine='xlrd')
                except:
                    # Last resort: use openpyxl in read-only mode
                    from openpyxl import load_workbook
                    wb = load_workbook(source_file, read_only=True, data_only=True)
                    
                    # Load Data sheet
                    ws_data = wb['Data']
                    data_values = list(ws_data.values)
                    data_cols = data_values[0]
                    self.data_df = pd.DataFrame(data_values[1:], columns=data_cols)
                    
                    # Load Layout sheet
                    ws_layout = wb['Layout']
                    layout_values = list(ws_layout.values)
                    layout_cols = layout_values[0]
                    self.layout_df = pd.DataFrame(layout_values[1:], columns=layout_cols)
                    
                    wb.close()
            else:
                raise
        
        print(f"Loaded {len(self.data_df)} rows from Data tab")
        print(f"Loaded {len(self.layout_df)} rows from Layout tab")
        
        
    def parse_layout(self):
        """
        Parse the Layout tab to extract question metadata and identify sub-questions.
        
        Returns:
            OrderedDict mapping Question IDs to their metadata, preserving Layout tab order
        """
        print("\nParsing layout information...")
        from collections import OrderedDict
        questions = OrderedDict()
        question_order = []  # Track order of first appearance
        
        for idx, row in self.layout_df.iterrows():
            question_id = row.get('Question ID', '')
            if pd.notna(question_id) and question_id != '':
                question_id = str(question_id).strip()
                column_label = row.get('Column Label', '')
                sub_question_text = row.get('Sub Question Text', '')
                
                if question_id not in questions:
                    question_order.append(question_id)
                    questions[question_id] = {
                        'column_label': column_label,
                        'answer_type': row.get('Answer Type', ''),
                        'question_text': row.get('Question Text', ''),
                        'group_text': row.get('Group Text', ''),
                        'position': row.get('Position', idx),  # Use row index if position not available
                        'report_as': row.get('Report As', ''),
                        'sub_questions': [],  # Track sub-question columns
                        'answers': [],
                        'order': len(question_order) - 1  # Store order for sorting
                    }
                
                # Track sub-questions (columns that belong to this question)
                # OR track answer options that have their own columns (multi-select)
                if pd.notna(sub_question_text) and sub_question_text != '':
                    questions[question_id]['sub_questions'].append({
                        'column_label': column_label,
                        'sub_question_text': sub_question_text,
                        'answer_text': row.get('Answer Text', '')
                    })
                elif pd.notna(column_label) and column_label != '' and column_label != questions[question_id]['column_label']:
                    # This is a multi-select option with its own column
                    answer_text = row.get('Answer Text', '')
                    if pd.notna(answer_text) and answer_text != '':
                        questions[question_id]['sub_questions'].append({
                            'column_label': column_label,
                            'sub_question_text': answer_text,  # Use answer text as sub-question text
                            'answer_text': answer_text
                        })
                
                # Collect answer options with values
                answer_text = row.get('Answer Text', '')
                answer_value = row.get('Answer Values...', '')  # Try the actual column name
                if pd.isna(answer_value):
                    answer_value = row.get('Answer Values', '')
                
                if pd.notna(answer_text) and answer_text != '':
                    questions[question_id]['answers'].append({
                        'text': answer_text,
                        'value': answer_value if pd.notna(answer_value) else ''
                    })
        
        print(f"Found {len(questions)} unique questions")
        return questions
    
    def get_answer_label_mapping(self, question_id):
        """
        Get mapping of numeric codes to answer labels for a question.
        Answer values are spread across multiple columns (N onwards) in the Layout tab.
        
        Args:
            question_id: Question ID to look up
            
        Returns:
            Dictionary mapping values to labels (e.g., {1: "1 = Alabama", 2: "2 = Alaska"})
        """
        matching_rows = self.layout_df[self.layout_df['Question ID'] == question_id]
        if matching_rows.empty:
            return {}
        
        mapping = {}
        
        # Find the "Answer Values..." column and all columns after it
        answer_values_col_idx = None
        for idx, col in enumerate(self.layout_df.columns):
            if 'Answer Values' in str(col):
                answer_values_col_idx = idx
                break
        
        if answer_values_col_idx is None:
            return {}
        
        # Process each row for this question
        for _, row in matching_rows.iterrows():
            answer_text = row.get('Answer Text', '')
            
            # Read answer values from column N onwards (all columns after Answer Values)
            # Each cell contains one answer value
            for col_idx in range(answer_values_col_idx, len(self.layout_df.columns)):
                col_name = self.layout_df.columns[col_idx]
                answer_value = row.get(col_name, '')
                
                if pd.notna(answer_value) and str(answer_value).strip() != '':
                    answer_value_str = str(answer_value).strip()
                    
                    # Check if answer_value already contains the format "value = text"
                    if '=' in answer_value_str:
                        parts = answer_value_str.split('=', 1)
                        if len(parts) == 2:
                            value_part = parts[0].strip()
                            label = answer_value_str  # Use the full string as label
                            
                            # Add string version
                            mapping[value_part] = label
                            
                            # Try to add numeric versions
                            try:
                                value_int = int(float(value_part))
                                mapping[value_int] = label
                                value_float = float(value_part)
                                mapping[value_float] = label
                            except (ValueError, TypeError):
                                pass
                    
                    # Case 2: Just the value, combine with Answer Text
                    elif pd.notna(answer_text) and answer_text != '':
                        label = f"{answer_value_str} = {answer_text}"
                        
                        # Add string version
                        mapping[answer_value_str] = label
                        
                        # Try to add numeric versions
                        try:
                            value_int = int(float(answer_value_str))
                            mapping[value_int] = label
                            value_float = float(answer_value_str)
                            mapping[value_float] = label
                        except (ValueError, TypeError):
                            pass
        
        return mapping
    
    def get_column_for_question(self, question_id):
        """
        Find the column name in Data tab for a given Question ID.
        
        Args:
            question_id: The question ID to search for
            
        Returns:
            Column name or None if not found
        """
        matching_rows = self.layout_df[self.layout_df['Question ID'] == question_id]
        if not matching_rows.empty:
            column_label = matching_rows.iloc[0]['Column Label']
            if column_label in self.data_df.columns:
                return column_label
        return None
    
    def is_numeric_input_question(self, question_id, question_col):
        """
        Determine if a question is a true numeric/percentage input question.
        Excludes questions with predefined answer options (single-select, multi-select, Likert scales).
        
        Args:
            question_id: Question ID to check in Layout tab
            question_col: Column name to check in Data tab
            
        Returns:
            Boolean indicating if the column is a numeric input question
        """
        matching_rows = self.layout_df[self.layout_df['Question ID'] == question_id]
        if matching_rows.empty:
            # No layout info, check if data is numeric
            valid_data = self.data_df[question_col].dropna()
            if len(valid_data) == 0:
                return False
            try:
                pd.to_numeric(valid_data, errors='raise')
                return True
            except:
                return False
        
        # Check Answer Type column (Column C) - most reliable indicator
        if 'Answer Type' in self.layout_df.columns:
            answer_types = matching_rows['Answer Type'].dropna().unique()
            for answer_type in answer_types:
                answer_type_str = str(answer_type).upper()
                # Percentage type questions should show means
                if 'PERCENTAGE' in answer_type_str or '%' in answer_type_str:
                    return True
                # Number input type questions should show means
                if 'NUMBER' in answer_type_str or 'NUMERIC' in answer_type_str:
                    return True
        
        # Check if this question has predefined answer values (Column N in Layout)
        # If Answer Values column has data, it's a single-select or multi-select question
        answer_value_columns = ['Answer Values', 'Answer Values...', 'AnswerValues', 'Answer_Values']
        
        for col_name in answer_value_columns:
            if col_name in self.layout_df.columns:
                # Check if column has any non-null, non-empty values
                values = matching_rows[col_name].dropna()
                has_answer_values = len(values) > 0 and values.astype(str).str.strip().ne('').any()
                
                if has_answer_values:
                    # Has predefined options (single-select, multi-select, or Likert)
                    # Show frequency distribution, not means
                    return False
                break  # Found the column, no need to check others
        
        # Check if data is numeric
        valid_data = self.data_df[question_col].dropna()
        if len(valid_data) == 0:
            return False
        
        try:
            pd.to_numeric(valid_data, errors='raise')
            return True
        except:
            return False
    
    def create_mean_crosstab(self, question_col, segment_columns, question_id=''):
        """
        Create a crosstab showing mean values for numeric/percentage questions.
        
        Args:
            question_col: Column name for the question
            segment_columns: Dictionary of segment names to column names
            question_id: Question ID for labeling
            
        Returns:
            DataFrame with mean values (Overall + all segments side-by-side)
        """
        # Filter out missing values for the question
        valid_data = self.data_df[[question_col] + list(segment_columns.values())].dropna(subset=[question_col])
        if len(valid_data) == 0:
            return None
        
        # Convert to numeric
        valid_data[question_col] = pd.to_numeric(valid_data[question_col], errors='coerce')
        valid_data = valid_data.dropna(subset=[question_col])
        
        if len(valid_data) == 0:
            return None
        
        # Create result row with empty first column for alignment
        row = {'': '', 'Metric': 'Mean'}
        
        # Overall mean and sample size (mean before n)
        row['Overall (Mean)'] = round(valid_data[question_col].mean(), 1)
        row['Overall (n)'] = len(valid_data)
        
        # Each segment
        for seg_name, seg_col in segment_columns.items():
            # Get data for this segment
            seg_data = valid_data[[question_col, seg_col]].dropna()
            if len(seg_data) == 0:
                continue
            
            # Get unique segment values
            seg_values = sorted(seg_data[seg_col].unique())
            
            # Get answer mapping for segment column
            seg_answer_mapping = self.get_answer_label_mapping(seg_name)
            
            for seg_value in seg_values:
                seg_subset = seg_data[seg_data[seg_col] == seg_value]
                if len(seg_subset) > 0:
                    mean_val = seg_subset[question_col].mean()
                    
                    # Apply mapping to segment value
                    seg_value_label = seg_answer_mapping.get(seg_value, seg_value)
                    
                    # Mean before n
                    row[f'{seg_name}_{seg_value_label} (Mean)'] = round(mean_val, 1)
                    row[f'{seg_name}_{seg_value_label} (n)'] = len(seg_subset)
        
        return pd.DataFrame([row])
    
    def create_banner_crosstab(self, question_col, segment_columns, question_id=''):
        """
        Create a banner-style crosstab with all segments side-by-side.
        For true numeric input questions, shows means instead of frequency distribution.
        For single-select questions (even with numeric codes), shows frequency distribution.
        
        Args:
            question_col: Column name for the question
            segment_columns: Dictionary of segment names to column names
            question_id: Question ID for labeling
            
        Returns:
            DataFrame with banner-style crosstab (Overall + all segments side-by-side)
        """
        # Check if this is a true numeric input question (not single-select with numeric codes)
        if self.is_numeric_input_question(question_id, question_col):
            return self.create_mean_crosstab(question_col, segment_columns, question_id)
        
        # Filter out missing values for the question
        valid_data = self.data_df[[question_col] + list(segment_columns.values())].dropna(subset=[question_col])
        if len(valid_data) == 0:
            return None
        
        # Get unique responses
        responses = sorted(valid_data[question_col].unique())
        
        # Get answer label mapping for this question
        answer_mapping = self.get_answer_label_mapping(question_id)
        
        # Create result dataframe
        result_rows = []
        
        for response in responses:
            # Use answer label if available, otherwise use raw response
            response_label = answer_mapping.get(response, response)
            # Add empty first column for alignment
            row = {'': '', 'Response': response_label}
            
            # Overall percentage and count (percentage before n)
            overall_count = (valid_data[question_col] == response).sum()
            overall_pct = (overall_count / len(valid_data) * 100)
            row['Overall (%)'] = round(overall_pct, 1)
            row['Overall (n)'] = overall_count
            
            # Each segment
            for seg_name, seg_col in segment_columns.items():
                # Get data for this segment
                seg_data = valid_data[[question_col, seg_col]].dropna()
                if len(seg_data) == 0:
                    continue
                
                # Get unique segment values
                seg_values = sorted(seg_data[seg_col].unique())
                
                # Get answer mapping for segment column as well
                seg_answer_mapping = self.get_answer_label_mapping(seg_name)
                
                for seg_value in seg_values:
                    seg_subset = seg_data[seg_data[seg_col] == seg_value]
                    count = (seg_subset[question_col] == response).sum()
                    pct = (count / len(seg_subset) * 100) if len(seg_subset) > 0 else 0
                    
                    # Apply mapping to segment value as well
                    seg_value_label = seg_answer_mapping.get(seg_value, seg_value)
                    
                    # Percentage before n
                    row[f'{seg_name}_{seg_value_label} (%)'] = round(pct, 1)
                    row[f'{seg_name}_{seg_value_label} (n)'] = count
            
            result_rows.append(row)
        
        # Add total row with empty first column
        total_row = {'': '', 'Response': 'Total'}
        total_row['Overall (%)'] = 100.0
        total_row['Overall (n)'] = len(valid_data)
        
        for seg_name, seg_col in segment_columns.items():
            seg_data = valid_data[[seg_col]].dropna()
            seg_values = sorted(seg_data[seg_col].unique())
            
            # Get answer mapping for segment column
            seg_answer_mapping = self.get_answer_label_mapping(seg_name)
            
            for seg_value in seg_values:
                count = (seg_data[seg_col] == seg_value).sum()
                # Apply mapping to segment value
                seg_value_label = seg_answer_mapping.get(seg_value, seg_value)
                # Percentage before n
                total_row[f'{seg_name}_{seg_value_label} (%)'] = 100.0
                total_row[f'{seg_name}_{seg_value_label} (n)'] = count
        
        result_rows.append(total_row)
        
        return pd.DataFrame(result_rows)
    
    def create_crosstab(self, question_col, segment_col=None, question_id='', segment_name='Overall'):
        """
        Create a crosstab with percentages and sample sizes.
        
        Args:
            question_col: Column name for the question
            segment_col: Column name for segmentation (None for overall)
            question_id: Question ID for labeling
            segment_name: Name of the segment
            
        Returns:
            DataFrame with crosstab results
        """
        # Filter out missing values
        if segment_col:
            valid_data = self.data_df[[question_col, segment_col]].dropna()
            if len(valid_data) == 0:
                return None
        else:
            valid_data = self.data_df[[question_col]].dropna()
            if len(valid_data) == 0:
                return None
        
        if segment_col:
            # Create crosstab with segmentation
            ct = pd.crosstab(
                valid_data[question_col],
                valid_data[segment_col],
                margins=True,
                margins_name='Total'
            )
            
            # Calculate percentages (column percentages)
            ct_pct = pd.crosstab(
                valid_data[question_col],
                valid_data[segment_col],
                normalize='columns',
                margins=True,
                margins_name='Total'
            ) * 100
            
            # Combine counts and percentages
            result = pd.DataFrame()
            for col in ct.columns:
                result[f'{col} (n)'] = ct[col]
                result[f'{col} (%)'] = ct_pct[col].round(1)
            
        else:
            # Overall crosstab (no segmentation)
            counts = valid_data[question_col].value_counts()
            percentages = (counts / len(valid_data) * 100).round(1)
            
            result = pd.DataFrame({
                'Response': counts.index,
                'Count (n)': counts.values,
                'Percentage (%)': percentages.values
            })
            
            # Add total row
            total_row = pd.DataFrame({
                'Response': ['Total'],
                'Count (n)': [len(valid_data)],
                'Percentage (%)': [100.0]
            })
            result = pd.concat([result, total_row], ignore_index=True)
        
        return result
    
    def create_combined_crosstab(self, question_id, q_info, segment_columns):
        """
        Create crosstab(s) for a question.
        - For array questions (sub-questions + answer values): Returns list of separate crosstabs
        - For other questions: Returns single combined crosstab
        
        Args:
            question_id: Question ID
            q_info: Question metadata dictionary
            segment_columns: Dictionary of segment names to column names
            
        Returns:
            Single DataFrame or list of DataFrames with crosstab(s)
        """
        has_sub_questions = q_info['sub_questions'] and len(q_info['sub_questions']) > 0
        has_answer_values = q_info['answers'] and len(q_info['answers']) > 0
        
        # Array question: sub-questions + answer values (e.g., A6, S7)
        # Create separate crosstab for each sub-question
        if has_sub_questions and has_answer_values:
            crosstabs = []
            
            # Check if all sub-questions have the same text (like S7 with "# of patients")
            # In that case, use answer_text to differentiate
            all_sub_texts = [sq['sub_question_text'] for sq in q_info['sub_questions']]
            use_answer_text = len(set(all_sub_texts)) == 1 and all_sub_texts[0] != ''
            
            for sub_q in q_info['sub_questions']:
                sub_col = sub_q['column_label']
                if sub_col not in self.data_df.columns:
                    continue
                
                # For questions like S7 where sub_question_text is the same for all,
                # use answer_text (tumor types) to differentiate
                # For questions like A6 where sub_question_text differs (product names),
                # use sub_question_text
                if use_answer_text:
                    answer_text = sub_q.get('answer_text', '')
                    display_text = answer_text if answer_text else sub_q['sub_question_text']
                else:
                    display_text = sub_q['sub_question_text']
                
                # Check if numeric input
                if self.is_numeric_input_question(question_id, sub_col):
                    ct = self.create_mean_crosstab(sub_col, segment_columns, question_id)
                else:
                    ct = self.create_banner_crosstab(sub_col, segment_columns, question_id)
                
                if ct is not None:
                    crosstabs.append({
                        'sub_question_text': display_text,
                        'crosstab': ct
                    })
            
            return crosstabs if crosstabs else None
        
        # Regular question with sub-questions (no answer values)
        # Combine all sub-questions into one table
        elif has_sub_questions:
            all_rows = []
            for sub_q in q_info['sub_questions']:
                sub_col = sub_q['column_label']
                if sub_col not in self.data_df.columns:
                    continue
                
                # Always use sub_question_text for combined tables
                sub_text = sub_q['sub_question_text']
                
                if self.is_numeric_input_question(question_id, sub_col):
                    ct = self.create_mean_crosstab(sub_col, segment_columns, question_id)
                    if ct is not None:
                        # Remove the empty first column and replace with Sub-Question
                        if '' in ct.columns:
                            ct = ct.drop(columns=[''])
                        ct.insert(0, 'Sub-Question', sub_text)
                        all_rows.append(ct)
                else:
                    ct = self.create_banner_crosstab(sub_col, segment_columns, question_id)
                    if ct is not None:
                        # Remove the empty first column and replace with Sub-Question
                        if '' in ct.columns:
                            ct = ct.drop(columns=[''])
                        ct.insert(0, 'Sub-Question', sub_text)
                        all_rows.append(ct)
            
            if not all_rows:
                return None
            return pd.concat(all_rows, ignore_index=True)
        
        # Single question - just create one crosstab
        else:
            question_col = self.get_column_for_question(question_id)
            if question_col:
                ct = self.create_banner_crosstab(question_col, segment_columns, question_id)
                return ct if ct is not None else None
            return None
    
    def generate_all_crosstabs(self):
        """
        Generate banner-style crosstabs for all questions, handling sub-questions properly.
        
        Returns:
            List of dictionaries containing crosstab results
        """
        questions = self.parse_layout()
        all_crosstabs = []
        
        # Get segment columns
        segment_columns = {}
        for seg_id in self.segment_questions:
            col = self.get_column_for_question(seg_id)
            if col:
                segment_columns[seg_id] = col
                print(f"Segment question {seg_id} found in column: {col}")
            else:
                print(f"Warning: Segment question {seg_id} not found in data")
        
        print(f"\nGenerating crosstabs for {len(questions)} questions...")
        
        for question_id, q_info in questions.items():
            # Skip segment questions themselves
            if question_id in self.segment_questions:
                continue
            
            print(f"Processing question: {question_id}")
            
            # Create crosstab(s) for the question
            result = self.create_combined_crosstab(question_id, q_info, segment_columns)
            
            if result is not None:
                # Check if it's a list of crosstabs (array question) or single crosstab
                if isinstance(result, list):
                    # Array question - add each sub-question's crosstab separately
                    for item in result:
                        all_crosstabs.append({
                            'question_id': question_id,
                            'question_text': q_info['question_text'],
                            'sub_question_text': item['sub_question_text'],
                            'crosstab': item['crosstab'],
                            'order': q_info.get('order', 999)
                        })
                else:
                    # Single crosstab
                    all_crosstabs.append({
                        'question_id': question_id,
                        'question_text': q_info['question_text'],
                        'crosstab': result,
                        'order': q_info.get('order', 999)
                    })
        
        print(f"\nGenerated {len(all_crosstabs)} crosstabs")
        return all_crosstabs
    
    def write_crosstabs_to_excel(self, crosstabs):
        """
        Write all banner-style crosstabs to a new sheet in the Excel workbook.
        
        Args:
            crosstabs: List of crosstab dictionaries
        """
        print("\nWriting crosstabs to Excel...")
        
        # Load the workbook
        wb = load_workbook(self.file_path)
        
        # Remove existing Crosstabs sheet if it exists
        if 'Crosstabs' in wb.sheetnames:
            del wb['Crosstabs']
        
        # Create new sheet
        ws = wb.create_sheet('Crosstabs', 0)  # Insert at beginning
        
        # Styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        subheader_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Sort crosstabs by original Layout tab order
        from itertools import groupby
        crosstabs_sorted = sorted(crosstabs, key=lambda x: (x.get('order', 999), x['question_id']))
        
        # Group by question to handle array questions
        from itertools import groupby
        
        for question_id, group in groupby(crosstabs_sorted, key=lambda x: x['question_id']):
            group_list = list(group)
            question_text = group_list[0]['question_text']
            
            # Question header
            max_col = 20  # Estimate, will adjust later
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"Question {question_id}: {question_text}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
            
            # Write each crosstab (for array questions, there will be multiple)
            for ct_info in group_list:
                # Sub-question header (for array questions)
                if ct_info.get('sub_question_text'):
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = f"  {ct_info['sub_question_text']}"
                    cell.font = subheader_font
                    cell.fill = subheader_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    current_row += 1
                
                df = ct_info['crosstab']
                
                # Write column headers
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                current_row += 1
                
                # Write data rows
                for _, row_data in df.iterrows():
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        
                        # Check column type
                        col_name = df.columns[col_idx-1] if col_idx > 0 else ''
                        is_percentage_col = '(%)' in col_name
                        is_mean_col = '(Mean)' in col_name
                        
                        # Format based on column type
                        if isinstance(value, (int, float)):
                            if is_percentage_col:
                                # Convert to decimal (divide by 100) and format as percentage
                                cell.value = value / 100
                                cell.number_format = '0.0%'
                            elif is_mean_col:
                                # Mean values as decimal
                                cell.value = value
                                cell.number_format = '0.0'
                            else:
                                cell.value = value
                        else:
                            cell.value = value
                        
                        cell.border = border
                    
                    current_row += 1
                
                # Add spacing after each sub-question's crosstab
                current_row += 1
            
            # Add extra spacing between questions
            current_row += 1
        
        # Adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(self.file_path)
        print(f"Crosstabs written to 'Crosstabs' sheet in {self.file_path}")
    
    def create_output_file(self):
        """Create a copy of the source file for output."""
        if self.output_file_path is None:
            # Auto-generate output filename
            base_name = os.path.splitext(self.source_file_path)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_file_path = f"{base_name}_with_Crosstabs_{timestamp}.xlsx"
        
        print(f"\nCreating output file...")
        print(f"Source: {self.source_file_path}")
        print(f"Output: {self.output_file_path}")
        
        # Copy the source file
        shutil.copy2(self.source_file_path, self.output_file_path)
        self.file_path = self.output_file_path
        print("✓ File copied successfully")
    
    def run(self):
        """Execute the complete crosstab generation process."""
        print("=" * 80)
        print("SURVEY DATA CROSSTAB GENERATOR")
        print("=" * 80)
        
        self.create_output_file()
        self.load_data()
        crosstabs = self.generate_all_crosstabs()
        self.write_crosstabs_to_excel(crosstabs)
        
        print("\n" + "=" * 80)
        print("CROSSTAB GENERATION COMPLETE!")
        print("=" * 80)
        print(f"\n✓ Output file created: {self.output_file_path}")
        print(f"✓ Original file unchanged: {self.source_file_path}")
        print(f"\nThe 'Crosstabs' tab has been added to the output workbook.")
        print(f"Total crosstabs generated: {len(crosstabs)}")
        print(f"\nSegmentation variables used:")
        for seg in self.segment_questions:
            print(f"  - {seg}")


if __name__ == "__main__":
    import sys
    
    # Get file path from command line or prompt user
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        print("\n" + "=" * 80)
        print("SURVEY DATA CROSSTAB GENERATOR")
        print("=" * 80)
        print("\nPlease provide the path to your Excel file.")
        print("You can either:")
        print("  1. Drag and drop the file here, or")
        print("  2. Type/paste the full file path")
        print("\nPress Enter to use default file, or provide path:")
        
        user_input = input("> ").strip().strip('"').strip("'")
        
        if user_input:
            file_path = user_input
        else:
            # Default file path
            file_path = r"Glioma_ATU_Survey_Wave_3_Data.xlsx"
            print(f"\nUsing default file: {file_path}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"\n❌ ERROR: File not found: {file_path}")
        print("Please check the file path and try again.")
        sys.exit(1)
    
    # Optional: Get output path from command line
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        # Create generator and run
        generator = CrosstabGenerator(file_path, output_path)
        generator.run()
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
